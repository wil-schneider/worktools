import pandas as pd
from openpyxl import load_workbook

# Specify your file path
xlsm_file = 'Sample-Spreadsheet-50000-rows.xlsm'

# Load workbook
wb = load_workbook(xlsm_file)

# Specify the sheet you want to parse
sheet_name = 'Sample-spreadsheet-file'

# Load sheet into a DataFrame
df = pd.read_excel(xlsm_file, sheet_name=sheet_name, engine='openpyxl')

# Write the data frame to a CSV file
df.to_csv('output.csv', index=False)

# Gather metadata
metadata = {
    'number_of_sheets': len(wb.sheetnames),
    'sheet_names': wb.sheetnames,
    'selected_sheet_name': sheet_name,
    'number_of_rows': df.shape[0],
    'number_of_columns': df.shape[1],
}

# Write metadata to a separate CSV file
pd.DataFrame([metadata]).to_csv('metadata.csv', index=False)
